"""
Tokinarc — apps/common/excel.py

Helper xuất Excel CHỨNG TỪ đẹp, dùng chung cho Báo giá / Hợp đồng / Phiếu xuất /
Phiếu nhập / Phiếu thu / Hóa đơn:
  - Tiêu đề + số/ngày chứng từ.
  - Khối thông tin đối tác (KH/NCC): tên, mã, MST, SĐT, địa chỉ.
  - Bảng dòng hàng: tiêu đề in đậm + nền, viền, giãn cột, format số/tiền.
  - Dòng tổng cộng.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_FLAME = 'E25822'
_THIN = Side(style='thin', color='CCCCCC')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill('solid', fgColor='F6E0D4')   # nền cam nhạt cho header bảng
_BOLD = Font(bold=True)
_MUTED = Font(color='777777')
_RIGHT = Alignment(horizontal='right')
_CENTER = Alignment(horizontal='center')

# columns: list[(header, width, kind)] — kind ∈ {'text','int','money'}


def customer_party(customer):
    """Khối thông tin khách hàng (SĐT lấy từ người liên hệ chính)."""
    addr = customer.address if isinstance(customer.address, dict) else {}
    address = ', '.join(p for p in [addr.get('street'), addr.get('district'),
                                    addr.get('city')] if p)
    pc = (customer.contacts.filter(is_primary=True).first()
          or customer.contacts.first())
    phone = pc.phone if pc else ''
    return [
        ('Tên KH:', customer.name),
        ('Mã KH:', customer.code),
        ('MST:', customer.tax_code or '—'),
        ('Điện thoại:', phone or '—'),
        ('Địa chỉ:', address or '—'),
    ]


def supplier_party(supplier):
    """Khối thông tin nhà cung cấp."""
    return [
        ('Tên NCC:', supplier.name),
        ('Mã NCC:', supplier.code),
        ('MST:', supplier.tax_code or '—'),
        ('Điện thoại:', supplier.phone or '—'),
        ('Địa chỉ:', supplier.address or '—'),
    ]


def _letterhead(ws, last_letter):
    """Đầu trang: logo (góc trái, nổi) + thông tin công ty (canh trái). Trả dòng kế tiếp."""
    from apps.common.company import COMPANY, logo_path
    lines = [
        (COMPANY['name'], Font(bold=True, size=13, color=_FLAME)),
        (f"ĐC: {COMPANY['address']}", Font(size=9)),
        (f"Showroom: {COMPANY['showroom']}", Font(size=9)),
        (f"MST: {COMPANY['tax_code']}     ĐT: {COMPANY['phone']}     Email: {COMPANY['email']}", Font(size=9)),
        (f"Website: {COMPANY['website']}", Font(size=9, color='1155CC')),
    ]
    r = 1
    for text, font in lines:
        ws.merge_cells(f'C{r}:{last_letter}{r}')
        c = ws.cell(r, 3, text)
        c.font = font
        c.alignment = Alignment(horizontal='left', vertical='center')
        r += 1
    lp = logo_path()
    if lp:
        try:
            img = XLImage(lp)
            h = 92
            img.width = int(h * (img.width / img.height)) if img.height else h
            img.height = h
            ws.add_image(img, 'A1')
        except Exception:
            pass
    return r + 1


def _signatures(ws, r, labels, ncols):
    r += 2
    seg = ncols / len(labels)
    for i, lab in enumerate(labels):
        col = int(i * seg) + 1
        end = max(col, int((i + 1) * seg))
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=end)
        c = ws.cell(r, col, lab); c.font = _BOLD; c.alignment = _CENTER
        ws.merge_cells(start_row=r + 1, start_column=col, end_row=r + 1, end_column=end)
        s = ws.cell(r + 1, col, '(Ký, ghi rõ họ tên)')
        s.font = Font(size=9, italic=True, color='888888'); s.alignment = _CENTER


def make_document_xlsx(*, sheet_title, doc_title, doc_code, doc_date=None,
                       party_label='KHÁCH HÀNG', party=None, meta=None,
                       columns=None, rows=None, total_label=None, total_value=None,
                       extra_totals=None, signatures=None, amount_words=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    # Cột STT đứng đầu (chứng từ luôn có)
    cols = [('STT', 6, 'int')] + list(columns) if columns else [('', 6, 'text')] * 4
    ncols = max(len(cols), 4)
    last = get_column_letter(ncols)

    # Đặt độ rộng cột trước (để khối letterhead canh đúng)
    for ci, (_h, w, _k) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    r = _letterhead(ws, last)

    ws.merge_cells(f'A{r}:{last}{r}')
    t = ws.cell(r, 1, doc_title)
    t.font = Font(bold=True, size=16, color=_FLAME)
    t.alignment = _CENTER
    r += 1
    sub = f"Số: {doc_code}" + (f"     Ngày: {doc_date}" if doc_date else '')
    ws.merge_cells(f'A{r}:{last}{r}')
    s = ws.cell(r, 1, sub); s.font = Font(size=10, color='666666'); s.alignment = _CENTER
    r += 2

    if party:
        ws.cell(r, 1, party_label).font = _BOLD
        r += 1
        for k, v in party:
            ws.cell(r, 1, k).font = _MUTED
            ws.merge_cells(f'B{r}:{last}{r}')
            ws.cell(r, 2, v)
            r += 1
        r += 1

    if meta:
        for k, v in meta:
            ws.cell(r, 1, k).font = _MUTED
            ws.merge_cells(f'B{r}:{last}{r}')
            ws.cell(r, 2, v)
            r += 1
        r += 1

    if columns:
        for ci, (h, _w, _kind) in enumerate(cols, start=1):
            cell = ws.cell(r, ci, h)
            cell.font = _BOLD
            cell.fill = _HEAD_FILL
            cell.border = _BORDER
            cell.alignment = _CENTER
        r += 1
        for idx, row in enumerate(rows or [], start=1):
            ws.cell(r, 1, idx).border = _BORDER
            ws.cell(r, 1).alignment = _CENTER
            for ci, (val, (_h, _w, kind)) in enumerate(zip(row, columns), start=2):
                cell = ws.cell(r, ci, val)
                cell.border = _BORDER
                if kind in ('money', 'int'):
                    cell.alignment = _RIGHT
                    if kind == 'money':
                        cell.number_format = '#,##0'
            r += 1
        for lab, val in (extra_totals or []):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols - 1)
            xc = ws.cell(r, 1, lab); xc.alignment = _RIGHT; xc.font = _MUTED
            xv = ws.cell(r, ncols, val); xv.number_format = '#,##0'; xv.alignment = _RIGHT
            r += 1
        if total_label is not None:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols - 1)
            lc = ws.cell(r, 1, total_label); lc.font = _BOLD; lc.alignment = _RIGHT
            tc = ws.cell(r, ncols, total_value)
            tc.font = Font(bold=True, color=_FLAME)
            tc.number_format = '#,##0'
            tc.alignment = _RIGHT
            r += 1
            if amount_words:
                ws.merge_cells(f'A{r}:{last}{r}')
                w = ws.cell(r, 1, f"Bằng chữ: {amount_words}")
                w.font = Font(italic=True, size=10)
                r += 1

    if signatures:
        _signatures(ws, r, signatures, ncols)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def xlsx_response(data: bytes, filename: str):
    from django.http import HttpResponse
    resp = HttpResponse(
        data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def style_table_sheet(ws, header_row=1, widths=None):
    """Làm đẹp 1 sheet bảng phẳng (cho file MISA): in đậm + nền header, viền, giãn cột."""
    max_col = ws.max_column
    for ci in range(1, max_col + 1):
        cell = ws.cell(header_row, ci)
        cell.font = _BOLD
        cell.fill = _HEAD_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER
    if widths:
        for ci, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=max_col):
        for cell in row:
            cell.border = _BORDER
