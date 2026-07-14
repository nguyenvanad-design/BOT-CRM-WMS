"""
Tokinarc V6 — apps/analytics/assistant.py

Trợ lý CRM NỘI BỘ (tách hẳn chatbot bán hàng cho khách). Trả lời câu hỏi nghiệp
vụ: doanh thu, công nợ, khách hàng — bằng cách:

  1. Hiểu câu hỏi → intent + tham số (LLM Gemini nếu có key, fallback từ khóa).
  2. Truy vấn DB thật (số liệu KHÔNG do LLM bịa — luôn tính từ Postgres).
  3. Soạn câu trả lời tiếng Việt từ số liệu thật.

MVP intent: revenue | customer_debt | top_customers | dormant_customers.
Bảo mật: chỉ manager/admin gọi (xem AssistantQueryView). KHÁC chatbot khách —
bot khách tuyệt đối không được chạm data này.
"""
from __future__ import annotations

import json
import os
import re
import urllib.error
import urllib.request
from pathlib import Path

from apps.analytics import api_client as apc


# ── Tiền VND ──────────────────────────────────────────────────────────────
def _num(x) -> int:
    """Ép về int an toàn — API trả *_vnd có thể là chuỗi Decimal ('950000.00')."""
    try:
        return int(float(x or 0))
    except (TypeError, ValueError):
        return 0


def _vnd(n) -> str:
    n = _num(n)
    try:
        n = int(n or 0)
    except (TypeError, ValueError):
        return '0 ₫'
    if abs(n) >= 1_000_000_000:
        return f"{n / 1e9:.2f}".rstrip('0').rstrip('.') + ' tỷ ₫'
    if abs(n) >= 1_000_000:
        return f"{round(n / 1e6):,}".replace(',', '.') + ' tr ₫'
    return f"{n:,}".replace(',', '.') + ' ₫'


# ── Tools (đọc QUA API — bot KHÔNG truy vấn DB trực tiếp) ────────────────────
def _api_or_err(user, path, query=None):
    """GET qua API; trả (data, None) hoặc (None, câu-lỗi) để tool trả thẳng."""
    try:
        return apc.get(user, path, query=query), None
    except apc.ApiError as e:
        if e.status == 403:
            return None, (f"Xin lỗi, bạn không có quyền xem dữ liệu này.")
        return None, f"Chưa lấy được dữ liệu (lỗi {e.status})."


def tool_revenue(user, period: str = 'month') -> str:
    d, err = _api_or_err(user, '/api/v1/analytics/revenue-summary/', {'period': period})
    if err:
        return err
    if not d['count']:
        return f"Doanh thu {d['label']}: chưa có đơn hàng nào."
    return (f"Doanh thu {d['label']}: **{_vnd(d['revenue_vnd'])}** từ {d['count']} đơn hàng "
            f"(đã thu {_vnd(d['paid_vnd'])}).")


def tool_customer_debt(user, name: str | None = None) -> str:
    d, err = _api_or_err(user, '/api/v1/analytics/customer-debt/',
                         {'customer': name} if name else None)
    if err:
        return err
    if d.get('found') is True:
        if d['debt_vnd'] <= 0:
            return f"**{d['name']}** hiện không còn công nợ (đã thanh toán đủ)."
        return (f"**{d['name']}** còn nợ **{_vnd(d['debt_vnd'])}** "
                f"(tổng đơn {_vnd(d['total_vnd'])}, đã trả {_vnd(d['paid_vnd'])}).")
    if d.get('found') is False:
        return f"Không tìm thấy khách hàng khớp \"{name}\"."
    rows = d.get('results') or []
    if not rows:
        return "Hiện không có công nợ phải thu."
    lines = '\n'.join(f"• {r['name']}: {_vnd(r['debt_vnd'])}" for r in rows)
    return f"Tổng công nợ phải thu (top {len(rows)}): **{_vnd(d['total_vnd'])}**\n{lines}"


def tool_top_customers(user, limit: int = 5) -> str:
    d, err = _api_or_err(user, '/api/v1/analytics/top-customers/', {'limit': limit})
    if err:
        return err
    if not d:
        return "Chưa có dữ liệu doanh số theo khách hàng."
    lines = '\n'.join(f"{i+1}. {r['name']}: {_vnd(r['revenue_vnd'])}" for i, r in enumerate(d))
    return f"Top {len(d)} khách hàng theo doanh số:\n{lines}"


def tool_dormant_customers(user, months: int = 3) -> str:
    d, err = _api_or_err(user, '/api/v1/analytics/dormant-customers/', {'months': months})
    if err:
        return err
    names = d.get('names') or []
    if not names:
        return f"Tất cả khách hàng đều có giao dịch trong {months} tháng gần đây."
    lines = '\n'.join(f"• {n}" for n in names)
    return (f"Có {d['count']} khách hàng không mua trong {months} tháng qua "
            f"(nguy cơ rời):\n{lines}")


def tool_reorder(user) -> str:
    """AI Reorder: đề nghị nhập hàng theo tốc độ bán + tồn khả dụng (đọc qua API)."""
    d, err = _api_or_err(user, '/api/v1/analytics/reorder-suggestions/')
    if err:
        return err
    if d['count'] == 0:
        return "Hiện không có mã nào cần nhập gấp — tồn đủ dùng. ✅"
    lines = []
    for r in d['results'][:12]:
        cover = f"đủ {r['days_cover']} ngày" if r['days_cover'] is not None else "chưa có lịch sử bán"
        lines.append(f"• {r['part_no']} {r['name'][:30]} — còn {r['available']} ({cover}) → nên nhập ~{r['suggest_qty']}")
    return (f"🛒 Đề nghị nhập hàng — {d['count']} mã sắp thiếu (đủ dùng < {d['lead_time_days']} ngày):\n"
            + "\n".join(lines))


def tool_slow_moving(user, days: int = 90) -> str:
    """AI Slow-moving: hàng bán chậm/chết — vốn đang chôn (đọc qua API)."""
    d, err = _api_or_err(user, '/api/v1/analytics/slow-moving/', {'days': days})
    if err:
        return err
    if d['count'] == 0:
        return f"Không có hàng bán chậm trong {days} ngày qua. 🎉"
    tied = f"{int(d['tied_value_vnd']):,}".replace(',', '.')
    lines = []
    for r in d['results'][:12]:
        idle = f"{r['days_idle']} ngày không xuất" if r['days_idle'] is not None else "chưa từng xuất"
        val = f"{int(r['value_vnd']):,}".replace(',', '.')
        lines.append(f"• {r['part_no']} {r['name'][:30]} — tồn {r['qty']}, vốn chôn {val}đ ({idle})")
    return (f"🐌 Hàng bán chậm/chết >{days} ngày — {d['count']} mã, vốn chôn {tied}đ:\n"
            + "\n".join(lines))


# ── Router: hiểu câu hỏi → intent + params ──────────────────────────────────
def _from_chatbot_env(var: str) -> str:
    """Đọc 1 biến từ chatbot/.env (dev, cùng máy)."""
    try:
        env = (Path(__file__).resolve().parents[3] / 'chatbot' / '.env').read_text(encoding='utf-8')
        m = re.search(rf'^{var}=(.*)$', env, re.M)
        return m.group(1).strip() if m else ''
    except OSError:
        return ''


def _gemini_key() -> str:
    return os.getenv('GEMINI_API_KEY', '') or _from_chatbot_env('GEMINI_API_KEY')


def _gemini_model() -> str:
    return (os.getenv('GEMINI_MODEL', '') or _from_chatbot_env('GEMINI_MODEL')
            or 'gemini-2.5-flash')


# ─── HYBRID PLANNER: Gemini function-calling (role-scoped) ───────────────────
# Mỗi "tool" = 1 intent (khớp dispatch trong answer()). Planner chỉ được khai báo
# các tool mà ROLE được phép → vừa chính xác hơn, vừa enforce quyền + ít token.
_FC_SYSTEM = (
    "Bạn là bộ định tuyến cho TRỢ LÝ NỘI BỘ Tokinarc (phân phối phụ tùng hàn). "
    "Đọc yêu cầu của nhân viên (kèm hội thoại trước nếu có) và GỌI ĐÚNG MỘT công cụ phù hợp, "
    "trích đủ tham số. Nếu câu mới là phần trả lời/bổ sung cho yêu cầu đang dở (vd cung cấp tên "
    "khách hàng cho báo giá), hãy gọi đúng công cụ của yêu cầu đó. Câu hỏi tra cứu kỹ thuật/sản "
    "phẩm/thông số chung → gọi lookup_doc. Hỏi CÁCH DÙNG PHẦN MỀM (vào menu nào, bấm gì, thao tác "
    "trên app — vd 'làm sao tạo báo giá', 'vào đâu xem công nợ', 'cách duyệt') → software_help (KHÁC "
    "procedure: procedure là kỹ thuật lắp/sửa SÚNG HÀN thật). Nếu không khớp công cụ nào (chào hỏi, "
    "ngoài phạm vi) → KHÔNG gọi công cụ nào.\n"
    "QUY TẮC GHI (create_*, wms_*):\n"
    "- LẦN ĐẦU nhận lệnh → confirm=false (bot XEM TRƯỚC, chưa ghi).\n"
    "- Nếu câu mới là SỬA/BỔ SUNG cho hành động ghi ĐANG CHỜ ở lượt trước (vd \"sửa công ty thành "
    "XYZ\", \"sđt là 090…\", \"đổi tên thành …\", \"thêm 5 x 002001\") → GỌI LẠI ĐÚNG công cụ đó với "
    "confirm=false, CẬP NHẬT tham số được sửa và GIỮ các tham số còn lại lấy từ bản xem trước trong "
    "hội thoại.\n"
    "- CHỈ đặt confirm=true KHI lượt trước đã có bản xem trước VÀ câu mới là xác nhận (ok/đồng ý/ghi "
    "đi); khi đó gọi LẠI đúng công cụ và GIỮ NGUYÊN mọi tham số đã chốt."
)


def _p(props, required=None):
    return {"type": "object", "properties": props, "required": required or []}


# Cờ xác nhận GHI: planner chỉ đặt true KHI nhân viên đã xem trước + nói ok/đồng ý.
_CONFIRM = {"type": "boolean",
            "description": "true CHỈ KHI nhân viên đã xem trước và xác nhận (ok/đồng ý/ghi đi) "
                           "trong hội thoại; lần ĐẦU nhận lệnh hoặc đang sửa = false"}


# (name khớp intent trong answer(); params khớp key intent.get(...) bên dispatch)
_FC_DECLS = [
    {"name": "revenue", "description": "Xem doanh thu/doanh số theo kỳ.",
     "parameters": _p({"period": {"type": "string", "description": "today | month | year"}})},
    {"name": "customer_debt", "description": "Công nợ phải thu của 1 khách hàng.",
     "parameters": _p({"customer_name": {"type": "string"}})},
    {"name": "top_customers", "description": "Top khách hàng theo doanh số.", "parameters": _p({})},
    {"name": "dormant_customers", "description": "Khách hàng lâu không mua (ngủ đông, nguy cơ rời).",
     "parameters": _p({"months": {"type": "integer", "description": "số tháng, mặc định 3"}})},
    {"name": "ceo_report", "description": "Báo cáo/tóm tắt điều hành toàn công ty.", "parameters": _p({})},
    {"name": "evaluate_plan", "description": "Đánh giá kế hoạch/pipeline/dự báo kinh doanh.", "parameters": _p({})},
    {"name": "reorder_suggestion", "description": "Đề nghị NHẬP HÀNG: mã sắp thiếu theo tốc độ bán.", "parameters": _p({})},
    {"name": "slow_moving", "description": "Hàng BÁN CHẬM/CHẾT: tồn lâu không xuất, vốn chôn.", "parameters": _p({})},
    {"name": "create_lead", "description": "Tạo lead (khách tiềm năng) mới.",
     "parameters": _p({"lead_name": {"type": "string"}, "company": {"type": "string"},
                       "phone": {"type": "string"}, "confirm": _CONFIRM}, ["lead_name"])},
    {"name": "create_quote", "description": "Lập báo giá nháp cho 1 khách hàng (đã là Customer).",
     "parameters": _p({"customer_name": {"type": "string"}, "confirm": _CONFIRM}, ["customer_name"])},
    {"name": "create_contract", "description": "Soạn hợp đồng từ 1 báo giá đã có.",
     "parameters": _p({"customer_name": {"type": "string"}, "quote_code": {"type": "string"},
                       "confirm": _CONFIRM})},
    {"name": "wms_inbound", "description": "Lập phiếu NHẬP KHO.",
     "parameters": _p({"confirm": _CONFIRM})},
    {"name": "wms_outbound", "description": "Lập phiếu XUẤT KHO/giao hàng.",
     "parameters": _p({"customer_name": {"type": "string"}, "confirm": _CONFIRM})},
    {"name": "customer_orders", "description": "Đơn hàng/lịch sử mua của 1 khách hàng.",
     "parameters": _p({"customer_name": {"type": "string"}})},
    {"name": "stock_lookup", "description": "Tra tồn kho 1 mã ở các kho.", "parameters": _p({})},
    {"name": "procedure", "description": "Hướng dẫn lắp đặt/sửa chữa/thông số kỹ thuật (liner, tip, torque...).", "parameters": _p({})},
    {"name": "compatibility", "description": "Phụ kiện đi kèm/tương thích với 1 sản phẩm.", "parameters": _p({})},
    {"name": "consumable_set", "description": "Bộ vật tư tiêu hao cho 1 model súng hàn.", "parameters": _p({})},
    {"name": "lookup_doc", "description": "Tra cứu tài liệu/sản phẩm/thông số Tokin chung.", "parameters": _p({})},
    {"name": "software_help", "description": "Hướng dẫn DÙNG PHẦN MỀM: vào menu nào, bấm gì, các bước "
     "thao tác trên app (khi nhân viên quên/chưa biết cách làm trên phần mềm). KHÁC procedure (procedure "
     "là kỹ thuật lắp/sửa SÚNG HÀN thật).", "parameters": _p({})},
]


def _fc_planner(question: str, role: str, history=None) -> dict | None:
    """Planner function-calling: Gemini chọn 1 tool (chỉ trong các tool role được phép)
    + trích tham số → trả intent dict cho dispatch. None nếu không gọi tool / lỗi (→ fallback)."""
    from apps.accounts.roles import can_use_intent
    key = _gemini_key()
    if not key:
        return None
    decls = [d for d in _FC_DECLS if can_use_intent(role, d['name'])]
    if not decls:
        return None
    model = _gemini_model()
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}"
    contents = []
    for h in (history or [])[-6:]:
        r = 'user' if (h.get('role') or '').lower() in ('user', 'human') else 'model'
        t = (h.get('text') or '').strip().replace("\n", " ")[:300]
        if t:
            contents.append({"role": r, "parts": [{"text": t}]})
    contents.append({"role": "user", "parts": [{"text": question}]})
    body = json.dumps({
        "system_instruction": {"parts": [{"text": _FC_SYSTEM}]},
        "contents": contents,
        "tools": [{"functionDeclarations": decls}],
        "toolConfig": {"functionCallingConfig": {"mode": "AUTO"}},
        "generationConfig": {"temperature": 0, "maxOutputTokens": 256,
                             "thinkingConfig": {"thinkingBudget": 0}},
    }).encode('utf-8')
    try:
        req = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=12) as resp:
            data = json.loads(resp.read())
        for p in data['candidates'][0]['content']['parts']:
            fc = p.get('functionCall')
            if fc and fc.get('name'):
                return {'intent': fc['name'], **(fc.get('args') or {})}
        return None
    except (urllib.error.URLError, KeyError, IndexError, ValueError, TimeoutError):
        return None


def _gemini_generate(system: str, prompt: str, max_tokens: int = 700) -> str | None:
    """Gọi Gemini sinh văn bản (system + prompt) → text. None nếu lỗi/không có key."""
    key = _gemini_key()
    if not key:
        return None
    model = _gemini_model()
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}"
    body = json.dumps({
        "system_instruction": {"parts": [{"text": system}]},
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.3, "maxOutputTokens": max_tokens,
                             "thinkingConfig": {"thinkingBudget": 0}},
    }).encode('utf-8')
    try:
        req = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
        return data['candidates'][0]['content']['parts'][0]['text'].strip()
    except (urllib.error.URLError, KeyError, IndexError, ValueError, TimeoutError):
        return None


def _keyword_intent(q: str) -> dict:
    ql = q.lower()
    if any(k in ql for k in ('tạo lead', 'tao lead', 'thêm lead', 'them lead', 'lead mới',
                             'khách tiềm năng', 'khach tiem nang', 'ghi lead')):
        return {'intent': 'create_lead'}
    # Hợp đồng trước báo giá: câu "soạn hợp đồng từ báo giá BG-x" chứa cả 2 từ khóa.
    if any(k in ql for k in ('hợp đồng', 'hop dong', 'soạn hợp đồng', 'lập hợp đồng')):
        return {'intent': 'create_contract'}
    if any(k in ql for k in ('báo giá', 'bao gia', 'tạo báo giá', 'lập báo giá', 'làm báo giá')):
        return {'intent': 'create_quote', 'items': _parse_items(q)}
    if any(k in ql for k in ('nhập kho', 'nhap kho', 'phiếu nhập', 'phieu nhap', 'nhận hàng')):
        return {'intent': 'wms_inbound'}
    if any(k in ql for k in ('xuất kho', 'xuat kho', 'phiếu xuất', 'phieu xuat', 'giao hàng')):
        return {'intent': 'wms_outbound'}
    if any(k in ql for k in ('báo cáo ceo', 'bao cao ceo', 'tóm tắt điều hành', 'tom tat dieu hanh',
                             'báo cáo điều hành', 'executive')):
        return {'intent': 'ceo_report'}
    if any(k in ql for k in ('đánh giá kế hoạch', 'danh gia ke hoach', 'pipeline', 'dự báo', 'du bao',
                             'forecast', 'kế hoạch kinh doanh')):
        return {'intent': 'evaluate_plan'}
    if any(k in ql for k in ('đơn của', 'don cua', 'đơn hàng của', 'don hang cua', 'mua gì', 'lịch sử mua')):
        return {'intent': 'customer_orders'}
    if any(k in ql for k in ('đề nghị nhập', 'de nghi nhap', 'nên nhập', 'nen nhap', 'cần nhập', 'can nhap',
                             'reorder', 'nhập thêm', 'nhap them', 'đặt thêm hàng', 'cần đặt hàng', 'sắp hết hàng gì')):
        return {'intent': 'reorder_suggestion'}
    if any(k in ql for k in ('bán chậm', 'ban cham', 'hàng chậm', 'hang cham', 'hàng ế', 'hang e',
                             'chậm luân chuyển', 'cham luan chuyen', 'tồn lâu', 'ton lau',
                             'slow moving', 'dead stock', 'hàng chết', 'hang chet')):
        return {'intent': 'slow_moving'}
    if any(k in ql for k in ('tồn', 'ton kho', 'còn hàng', 'con hang', 'kho còn', 'số lượng tồn')):
        return {'intent': 'stock_lookup'}
    # Hướng dẫn dùng PHẦN MỀM (đặt TRƯỚC procedure vì cùng có "hướng dẫn/cách").
    if any(k in ql for k in ('làm sao', 'lam sao', 'làm thế nào', 'lam the nao', 'vào đâu', 'vao dau',
                             'ở đâu', 'o dau', 'vào menu', 'bấm vào đâu', 'bam vao dau', 'thao tác trên',
                             'dùng phần mềm', 'dung phan mem', 'sử dụng phần mềm', 'su dung phan mem',
                             'hướng dẫn sử dụng', 'huong dan su dung', 'cách tạo', 'cach tao', 'cách lập',
                             'cach lap', 'cách duyệt', 'cach duyet', 'cách xem', 'cach xem')):
        return {'intent': 'software_help'}
    if any(k in ql for k in ('cách thay', 'cach thay', 'cách lắp', 'cach lap', 'quy trình', 'quy trinh',
                             'lắp đặt', 'lap dat', 'hướng dẫn', 'huong dan', 'torque', 'lực vặn', 'luc van',
                             'cắt liner', 'cat liner', 'chiều dài liner', 'inner tube', 'độ nhô', 'do nho',
                             'sửa', 'sua chua', 'khắc phục', 'khac phuc')):
        return {'intent': 'procedure'}
    if any(k in ql for k in ('đi kèm', 'di kem', 'dùng chung', 'dung chung', 'tương thích', 'tuong thich',
                             'đi với', 'di voi', 'companion', 'lắp được với', 'lap duoc voi')):
        return {'intent': 'compatibility'}
    if any(k in ql for k in ('bộ tiêu hao', 'bo tieu hao', 'vật tư cho', 'vat tu cho', 'set linh kiện',
                             'bộ linh kiện', 'bo linh kien', 'bộ vật tư', 'consumable')):
        return {'intent': 'consumable_set'}
    if any(k in ql for k in ('công nợ', 'cong no', 'còn nợ', 'no bao nhieu', 'nợ', 'phải thu')):
        # Thử bắt tên KH sau từ khóa (đơn giản): cụm chữ hoa hoặc sau "của"
        m = re.search(r'(?:của|cua)\s+([\w\sÀ-ỹ]+?)(?:\s+còn|\s+nợ|\?|$)', q, re.I)
        return {'intent': 'customer_debt', 'customer_name': (m.group(1).strip() if m else '')}
    if any(k in ql for k in ('chưa mua', 'chua mua', 'ngủ đông', 'ngu dong', 'nguy cơ rời', 'rời bỏ')):
        mm = re.search(r'(\d+)\s*tháng', ql)
        return {'intent': 'dormant_customers', 'months': int(mm.group(1)) if mm else 3}
    if any(k in ql for k in ('top', 'nhiều nhất', 'nhieu nhat', 'lớn nhất', 'đóng góp')):
        return {'intent': 'top_customers'}
    if any(k in ql for k in ('doanh thu', 'doanh số', 'doanh so', 'revenue')):
        if any(k in ql for k in ('hôm nay', 'hom nay', 'today')):
            period = 'today'
        elif any(k in ql for k in ('năm', 'nam nay', 'year')):
            period = 'year'
        else:
            period = 'month'
        return {'intent': 'revenue', 'period': period}
    if any(k in ql for k in ('tra cứu', 'tra cuu', 'thông tin', 'thong tin', 'thông số',
                             'thong so', 'spec', 'tài liệu', 'tai lieu', 'phụ tùng', 'phu tung',
                             'súng hàn', 'sung han')):
        return {'intent': 'lookup_doc'}
    return {'intent': 'unknown'}


def _detect_customer(user, q: str) -> str:
    """Đối chiếu câu hỏi với tên KH QUA API (bắt tên dù LLM/regex bỏ sót; đã lọc theo quyền)."""
    ql = q.lower()
    rows = apc.results(_safe_get(user, '/api/v1/crm/customers/',
                                 {'page_size': 500, 'ordering': 'name'}))
    for c in rows:
        n = c.get('name') or ''
        if n and n.lower() in ql:
            return n
    return ''


def _parse_items(q: str) -> list[dict]:
    """Trích (mã phụ tùng, số lượng) từ câu tự do — fallback khi không có LLM.

    Bắt mẫu: "5 x 001002", "10 cái 002001", "001002 x3", "2 001003".
    """
    items: list[dict] = []
    seen = set()
    # qty trước mã: "5 x 001002" / "10 cái 002001" / "2 001003"
    for m in re.finditer(r'(\d+)\s*(?:x|cái|cai|chiếc|chiec|pcs|\*)?\s*([0-9]{4,}[A-Za-z0-9\-]*)', q):
        qty, pn = int(m.group(1)), m.group(2)
        if pn not in seen:
            items.append({'part_no': pn, 'qty': max(1, qty)}); seen.add(pn)
    # mã trước qty: "001002 x3"
    for m in re.finditer(r'([0-9]{4,}[A-Za-z0-9\-]*)\s*(?:x|\*)\s*(\d+)', q):
        pn, qty = m.group(1), int(m.group(2))
        if pn not in seen:
            items.append({'part_no': pn, 'qty': max(1, qty)}); seen.add(pn)
    return items


# ── Tool điều hành (đọc qua API) ────────────────────────────────────────────
def tool_ceo_report(user) -> str:
    """Báo cáo điều hành cho CEO (tóm tắt toàn phòng ban, số liệu thật qua API)."""
    return executive_summary(user)['summary']


def tool_evaluate_plan(user) -> str:
    """Đánh giá kế hoạch kinh doanh từ pipeline forecast (đọc qua API)."""
    rows, err = _api_or_err(user, '/api/v1/analytics/forecast/pipeline/')
    if err:
        return err
    if not rows:
        return "Chưa có cơ hội nào trong pipeline để đánh giá kế hoạch."
    total_w = sum(float(r['weighted_vnd'] or 0) for r in rows)
    total_cnt = sum(int(r.get('count', 0) or 0) for r in rows)
    lines = [f"• {r['stage']}: {r.get('count', 0)} cơ hội — dự báo có trọng số "
             f"{_vnd(r['weighted_vnd'])}" for r in rows]
    body = '\n'.join(lines)
    return (f"**Đánh giá kế hoạch (pipeline)**\n"
            f"{total_cnt} cơ hội đang mở; **dự báo doanh thu có trọng số "
            f"(weighted): {_vnd(total_w)}**.\n{body}\n"
            f"_Weighted = giá trị cơ hội × xác suất theo giai đoạn._")


# ── Tool ghi: làm báo giá nháp ──────────────────────────────────────────────
def _parse_phone(q: str) -> str:
    m = re.search(r'\b0\d{8,10}\b', q.replace('.', '').replace(' ', ''))
    return m.group(0) if m else ''


# ── Phân giải QUA API (không ORM): khách hàng + phụ tùng ────────────────────
def _api_customer(user, name: str):
    """Tìm khách hàng QUA API (endpoint tự lọc theo quyền sở hữu). Trả dict KH hoặc None."""
    name = (name or '').strip()
    if not name:
        return None
    try:
        rows = apc.results(apc.get(user, '/api/v1/crm/customers/', query={'search': name}))
    except apc.ApiError:
        return None
    if not rows:
        return None
    low = name.lower()
    for c in rows:                                   # ưu tiên khớp MÃ chính xác
        if (c.get('code') or '').lower() == low:
            return c
    for c in rows:                                   # rồi tên/mã chứa chuỗi
        if low in (c.get('name') or '').lower() or low in (c.get('code') or '').lower():
            return c
    return rows[0]


def _api_part(user, part_no: str):
    """Lấy 1 phụ tùng QUA API theo mã (PK = tokin_part_no). None nếu không có."""
    pn = (part_no or '').strip()
    if not pn:
        return None
    try:
        return apc.get(user, f'/api/v1/catalog/parts/{pn}/')
    except apc.ApiError:
        return None


def _confirm_hint() -> str:
    """Nhắc xác nhận trước khi GHI (mọi phòng ban: ok mới ghi, sai thì sửa)."""
    return "\n\n→ Đúng chưa? Gõ **ok** để ghi · hoặc **sửa** (vd \"tên là …\", \"sđt 090…\", \"5 x 002001\")."


def _preview(title: str, body: str = '') -> str:
    """Khung 'xem trước' dùng chung cho mọi tool ghi → nhân sự ok mới ghi."""
    head = f"📝 **Xem trước — {title}**"
    return (f"{head}:\n{body}" if body else head) + _confirm_hint()


def _part_rows(found, missing) -> tuple[str, str]:
    """(rows, note) cho danh sách dòng hàng phiếu nhập/xuất."""
    rows = '\n'.join(f"• {n} (`{pn}`) × {qty}" for pn, n, qty, _ in found)
    note = f"\n⚠️ Không thấy mã: {', '.join(missing)}" if missing else ""
    return rows, note


def tool_create_lead(user, name: str, company: str = '', phone: str = '',
                     source: str = 'chatbot', confirm: bool = False) -> str:
    """Tạo LEAD (khách tiềm năng) QUA API, owner = người dùng. Xem trước → ok mới ghi."""
    name = (name or '').strip()
    if not name:
        return ('Cần tên khách tiềm năng. VD: "tạo lead Nguyễn Văn A, công ty ABC, 0901234567".')
    company = (company or '').strip()
    phone = (phone or '').strip()
    if not confirm:
        return _preview("sắp tạo lead",
                        f"• Tên: **{name}**\n• Công ty: {company or '—'}\n• SĐT: {phone or '—'}")
    try:
        lead = apc.post(user, '/api/v1/crm/leads/', {
            'name': name, 'company': company, 'phone': phone, 'source': source or 'chatbot'})
    except apc.ApiError as e:
        return f"Chưa ghi được lead: {e.detail}"
    extra = [x for x in (lead.get('company'), lead.get('phone')) if x]
    tail = f" ({', '.join(extra)})" if extra else ""
    return (f"✅ Đã ghi **lead: {lead['name']}**{tail} vào CRM (trạng thái Mới). "
            f"Vào menu Leads để theo dõi / chuyển thành khách hàng.")


def tool_create_quote(user, customer_name: str, items: list[dict], confirm: bool = False) -> str:
    """Tạo BÁO GIÁ NHÁP (draft) QUA API, lấy giá từ catalog. Xem trước → ok mới ghi."""
    if not customer_name:
        return "Cần cho biết **tên khách hàng** để lập báo giá. VD: \"làm báo giá cho Công ty ABC: 5 x 001002\"."
    cust = _api_customer(user, customer_name)
    if not cust:
        # Có thể là LEAD chưa chuyển thành khách hàng → gợi chuyển (báo giá cần Customer).
        try:
            leads = apc.results(apc.get(user, '/api/v1/crm/leads/',
                                        query={'search': customer_name.strip()}))
        except apc.ApiError:
            leads = []
        low = customer_name.strip().lower()
        lead = next((l for l in leads
                     if (l.get('status') in ('new', 'contacted', 'qualified'))
                     and (low in (l.get('name') or '').lower()
                          or low in (l.get('company') or '').lower())), None)
        if lead:
            return (f"\"{lead['name']}\" đang là **lead** (chưa phải khách hàng) → cần **chuyển thành khách hàng** "
                    f"trước khi báo giá. Vào **Leads → {lead['name']} → Chuyển + Tạo cơ hội**, rồi quay lại nói "
                    f"\"làm báo giá cho {lead.get('company') or lead['name']}\".")
        return (f"Không tìm thấy khách hàng khớp \"{customer_name}\" (trong phạm vi của bạn). "
                f"Kiểm tra lại tên/mã KH.")
    if not items:
        return (f"Đã xác định KH **{cust['name']}** nhưng chưa có dòng hàng. "
                f"Nêu mã phụ tùng + số lượng, VD: \"5 x 001002, 10 cái 002001\".")

    # Phân giải mã + giá TRƯỚC qua API (để xem trước); chưa ghi gì.
    found, missing = [], []
    for it in items:
        pn = str(it.get('part_no', '')).strip()
        qty = max(1, int(it.get('qty', 1) or 1))
        if not pn:
            continue
        part = _api_part(user, pn)
        if not part:
            missing.append(pn); continue
        found.append((pn, part.get('display_name_vi') or pn, qty, int(part.get('price_vnd') or 0)))

    if not found:
        return (f"Không tạo được báo giá: không tìm thấy mã phụ tùng nào hợp lệ "
                f"({', '.join(missing)}).")

    rows = '\n'.join(f"• {n} (`{pn}`) × {qty} = {_vnd(p * qty)}" for pn, n, qty, p in found)
    total = sum(p * qty for pn, n, qty, p in found)
    note = f"\n⚠️ Không thấy mã: {', '.join(missing)}" if missing else ""
    if not confirm:
        return _preview(f"sắp tạo báo giá cho {cust['name']}",
                        f"{rows}\n**Tổng tạm: {_vnd(total)}**{note}")

    # ĐÃ XÁC NHẬN → ghi QUA API (serializer tự tạo lines + tính total; view tự sinh
    # code + định tuyến duyệt + notify manager — không cần bot làm tay).
    try:
        quote = apc.post(user, '/api/v1/crm/quotes/', {
            'customer': cust['id'],
            'lines': [{'part_no': pn, 'part_name': n, 'qty': qty, 'unit_price_vnd': p}
                      for pn, n, qty, p in found],
        })
    except apc.ApiError as e:
        return f"Chưa tạo được báo giá: {e.detail}"

    # Câu chốt phản ánh ĐÚNG trạng thái sau khi ghi: view định tuyến duyệt theo chiết
    # khấu — báo giá bot lập (0% CK) thường TỰ DUYỆT (trong hạn mức sale), không phải Nháp.
    if quote.get('status') == 'approved':
        tail = ("Báo giá **đã tự duyệt** (chiết khấu trong hạn mức) — vào màn Báo giá "
                "bấm **Tạo đơn** để lên đơn bán.")
    else:
        lvl = " (giá trị lớn — sẽ cần CEO duyệt cấp 2)" if quote.get('requires_l2') else ""
        tail = f"Báo giá đang chờ **trình duyệt**{lvl} — vào màn Báo giá để gửi & duyệt."
    return (f"✅ Đã tạo **báo giá {quote['code']}** cho **{cust['name']}**:\n{rows}\n"
            f"**Tổng: {_vnd(quote.get('total_vnd'))}**{note}\n{tail}")


def tool_create_contract(user, customer_name: str, quote_code: str, confirm: bool = False) -> str:
    """Soạn HỢP ĐỒNG NHÁP QUA API: ưu tiên từ báo giá đã duyệt (quote_code), hoặc cho 1 KH."""
    # Từ báo giá đã duyệt
    if quote_code:
        rows = apc.results(_safe_get(user, '/api/v1/crm/quotes/', {'search': quote_code}))
        quote = next((q for q in rows if (q.get('code') or '').lower() == quote_code.lower()), None)
        if not quote:
            return f"Không tìm thấy báo giá **{quote_code}** (trong phạm vi của bạn)."
        if quote.get('status') not in ('approved', 'converted'):
            return (f"Báo giá {quote['code']} đang *{quote.get('status_display')}* — "
                    f"phải **đã duyệt** mới soạn hợp đồng được.")
        if not confirm:
            return _preview(f"sắp soạn hợp đồng từ báo giá **{quote['code']}** "
                            f"cho **{quote.get('customer_name')}**, giá trị **{_vnd(quote.get('total_vnd'))}**")
        try:
            ct = apc.post(user, '/api/v1/crm/contracts/', {
                'customer': quote['customer'], 'quote': quote['id'],
                'value_vnd': quote.get('total_vnd'),
                'title': f"Hợp đồng theo báo giá {quote['code']}"})
        except apc.ApiError as e:
            return f"Chưa soạn được hợp đồng: {e.detail}"
        return (f"✅ Đã soạn **hợp đồng nháp {ct['code']}** cho **{quote.get('customer_name')}** "
                f"từ báo giá {quote['code']}, giá trị **{_vnd(ct.get('value_vnd'))}**.\n"
                f"Vào màn Hợp đồng để bổ sung điều khoản & trình duyệt.")

    # Tạo trực tiếp cho 1 KH (chưa có giá trị)
    if not customer_name:
        return ("Cần **mã báo giá đã duyệt** (VD: BG-0007) hoặc **tên khách hàng** "
                "để soạn hợp đồng. VD: \"soạn hợp đồng từ báo giá BG-0007\".")
    cust = _api_customer(user, customer_name)
    if not cust:
        return f"Không tìm thấy khách hàng khớp \"{customer_name}\" (trong phạm vi của bạn)."
    if not confirm:
        return _preview(f"sắp soạn hợp đồng cho **{cust['name']}** (chưa có giá trị)")
    try:
        ct = apc.post(user, '/api/v1/crm/contracts/', {
            'customer': cust['id'], 'title': f"Hợp đồng — {cust['name']}"})
    except apc.ApiError as e:
        return f"Chưa soạn được hợp đồng: {e.detail}"
    return (f"✅ Đã soạn **hợp đồng nháp {ct['code']}** cho **{cust['name']}** (chưa có giá trị).\n"
            f"Vào màn Hợp đồng để nhập giá trị, điều khoản & trình duyệt.")


# ── Tool ghi: phiếu nhập / xuất kho (QUA API) ───────────────────────────────
def _api_default_warehouse(user):
    """Kho mặc định QUA API: nếu chỉ 1 kho active → dùng luôn; nhiều kho → lấy is_default."""
    whs = apc.results(_safe_get(user, '/api/v1/wms/warehouses/'))
    actives = [w for w in whs if w.get('is_active')]
    if len(actives) == 1:
        return actives[0]
    return next((w for w in actives if w.get('is_default')), None)


def _api_part_lines(user, items: list[dict]):
    """[{part_no, qty}] → (found[(pn,name,qty,None)], missing[pn]) — resolve QUA API."""
    found, missing = [], []
    for it in items or []:
        pn = str(it.get('part_no', '')).strip()
        qty = max(1, int(it.get('qty', 1) or 1))
        if not pn:
            continue
        part = _api_part(user, pn)
        if part:
            found.append((pn, part.get('display_name_vi') or pn, qty, None))
        else:
            missing.append(pn)
    return found, missing


def tool_wms_inbound(user, items: list[dict], confirm: bool = False) -> str:
    """Lập PHIẾU NHẬP KHO nháp QUA API. Xem trước → ok mới ghi."""
    wh = _api_default_warehouse(user)
    if not wh:
        return "Hệ thống có nhiều kho — vui lòng lập phiếu nhập trên màn Nhập kho và chọn kho cụ thể."
    if not items:
        return "Cần mã phụ tùng + số lượng để lập phiếu nhập. VD: \"nhập kho 100 x 001002\"."
    found, missing = _api_part_lines(user, items)
    if not found:
        return f"Không lập được phiếu nhập: không có mã phụ tùng hợp lệ ({', '.join(missing)})."
    rows, note = _part_rows(found, missing)
    if not confirm:
        return _preview(f"sắp lập phiếu NHẬP kho (kho {wh['code']})", f"{rows}{note}")
    try:
        order = apc.post(user, '/api/v1/wms/inbound/', {
            'warehouse': wh['id'],
            'lines': [{'part': pn, 'qty_expected': qty} for pn, _n, qty, _ in found]})
    except apc.ApiError as e:
        return f"Chưa lập được phiếu nhập: {e.detail}"
    return (f"✅ Đã lập **phiếu nhập kho nháp {order['code']}** (kho {wh['code']}):\n{rows}{note}\n"
            f"Vào màn Nhập kho để xác nhận nhận hàng (cộng tồn).")


def tool_wms_outbound(user, items: list[dict], customer_name: str = '', confirm: bool = False) -> str:
    """Lập PHIẾU XUẤT KHO nháp QUA API. Xem trước → ok mới ghi."""
    wh = _api_default_warehouse(user)
    if not wh:
        return "Hệ thống có nhiều kho — vui lòng lập phiếu xuất trên màn Xuất kho và chọn kho cụ thể."
    if not items:
        return "Cần mã phụ tùng + số lượng để lập phiếu xuất. VD: \"xuất kho 20 x 001002\"."
    found, missing = _api_part_lines(user, items)
    if not found:
        return f"Không lập được phiếu xuất: không có mã phụ tùng hợp lệ ({', '.join(missing)})."
    cust = _api_customer(user, customer_name) if customer_name else None
    rows, note = _part_rows(found, missing)
    who = f" cho **{cust['name']}**" if cust else ""
    if not confirm:
        return _preview(f"sắp lập phiếu XUẤT kho (kho {wh['code']}){who}", f"{rows}{note}")
    try:
        order = apc.post(user, '/api/v1/wms/outbound/', {
            'warehouse': wh['id'], 'customer': cust['id'] if cust else None,
            'lines': [{'part': pn, 'qty_ordered': qty} for pn, _n, qty, _ in found]})
    except apc.ApiError as e:
        return f"Chưa lập được phiếu xuất: {e.detail}"
    return (f"✅ Đã lập **phiếu xuất kho nháp {order['code']}** (kho {wh['code']}){who}:\n{rows}{note}\n"
            f"Vào màn Xuất kho để soạn hàng (pick) & giao.")


# ── Tool đọc: tra cứu tài liệu/sản phẩm Tokin (catalog) ─────────────────────
_SOFTWARE_GUIDE = (
    "Bạn là trợ lý HƯỚNG DẪN SỬ DỤNG phần mềm Tokinarc (ERP phân phối phụ tùng hàn của AUTOSS) cho "
    "NHÂN VIÊN NỘI BỘ quên/chưa biết thao tác. Trả lời NGẮN, theo TỪNG BƯỚC, chỉ rõ VÀO MENU NÀO / "
    "BẤM GÌ. Tiếng Việt thân thiện. Hỏi ngoài phần mềm → nói chưa rõ.\n\n"
    "MENU (mỗi vai trò thấy khu của mình):\n"
    "• CRM (sale/quản lý KD): Dashboard · Khách hàng · Leads · Cơ hội (nút Bảng/Kanban) · Báo giá · "
    "Đơn bán · Hợp đồng · Hóa đơn(MISA) · Công nợ · Nhật ký của tôi · Visit Report · Hoạt động · "
    "Service Ticket · Bảo hành · Trả hàng(RMA) · Sản phẩm.\n"
    "• WMS (NV kho/QL kho): Dashboard kho · Mua hàng(Đơn mua, Nhà cung cấp) · Nhập kho · Xuất kho · "
    "Tồn kho · Truy xuất(Serial/Lô) · Lịch sử kho · Kiểm kê & Tra cứu · Kho & vị trí · Bản đồ kho.\n"
    "• CEO (giám đốc/quản lý): Cần duyệt · Dashboard điều hành · AI Summary · Doanh thu · Công nợ · "
    "Forecast · Tồn kho · Tuổi tồn & Hàng chậm.\n"
    "• Admin: Người dùng & quyền. Góc trên phải mọi màn: 'Tài khoản của tôi' (đổi mật khẩu), 'Đăng xuất'.\n\n"
    "LUỒNG CHÍNH:\n"
    "• Bán hàng: Leads → mở lead bấm 'Chuyển + Tạo cơ hội' (thành Khách hàng + Cơ hội) → Báo giá "
    "(Tạo BG, nhập mã+SL, giá tự gợi theo phân khúc KH) → gửi & trình duyệt → quản lý/CEO vào 'Cần "
    "duyệt' duyệt → bấm 'Tạo đơn' → Đơn bán → kho vào Xuất kho (soạn/pick → Giao) → Hóa đơn(MISA) → "
    "Công nợ thu tiền.\n"
    "• Mua/nhập: Mua hàng → Đơn mua (Tạo, chọn NCC) → quản lý/CEO duyệt → 'Hàng đang về' theo dõi → "
    "hàng tới: Nhập kho (tạo phiếu hoặc từ đơn mua → quét/nhận → Xác nhận để cộng tồn).\n"
    "• Kiểm kê: Kiểm kê & Tra cứu → tab Kiểm kê → 'Phiên mới' → quét đếm (mã+ô+số) → xem chênh lệch → "
    "QL kho bấm 'Áp dụng' để chỉnh tồn. Tab Tra cứu: quét/nhập mã xem SP + tồn.\n"
    "• Đổi mật khẩu: bấm TÊN MÌNH góc trên phải → 'Tài khoản của tôi' → nhập MK hiện tại + MK mới.\n"
    "• Báo giá cho khách đang là LEAD: phải Chuyển lead → khách hàng trước.\n"
    "• Khung chat 'Trợ lý nội bộ' (dưới màn): gõ thẳng lệnh để LÀM nhanh — vd 'tạo lead ...', "
    "'làm báo giá cho ... 5 x 001002', 'nên nhập hàng gì', 'doanh thu tháng này' (tùy quyền)."
)


def tool_software_help(question: str, role: str = '') -> str:
    """Hướng dẫn DÙNG PHẦN MỀM (vào menu nào, bấm gì) cho nhân viên quên/chưa biết."""
    role_hint = f"(Người hỏi vai trò: {role}) " if role else ""
    out = _gemini_generate(_SOFTWARE_GUIDE, role_hint + "Câu hỏi: " + question)
    if out:
        return out
    return ("Em hướng dẫn dùng phần mềm. Vài nơi hay dùng:\n"
            "• **Báo giá**: menu Báo giá → Tạo BG (nhập mã + số lượng).\n"
            "• **Lead → khách**: Leads → mở lead → 'Chuyển + Tạo cơ hội'.\n"
            "• **Nhập/Xuất kho**: menu Nhập kho / Xuất kho.\n"
            "• **Kiểm kê**: Kiểm kê & Tra cứu → Phiên mới → quét đếm → Áp dụng.\n"
            "• **Đổi mật khẩu**: bấm tên mình góc phải → Tài khoản của tôi.\n"
            "Anh/chị muốn hướng dẫn thao tác cụ thể nào ạ?")


def _safe_get(user, path, query=None):
    """GET qua API, trả data hoặc None nếu lỗi (dùng cho tra cứu không cần báo lỗi chi tiết)."""
    try:
        return apc.get(user, path, query=query)
    except apc.ApiError:
        return None


def tool_lookup_doc(user, question: str) -> str:
    """Tra cứu phụ tùng/súng hàn Tokin từ catalog (mã hoặc tên) — QUA API."""
    codes = re.findall(r'\b[0-9]{4,}[A-Za-z0-9\-]*\b', question)
    part = None
    for cpn in codes:
        part = _api_part(user, cpn)
        if part:
            break
    if not part:
        # tìm theo tên: lấy cụm chữ có nghĩa cuối câu → search API → lấy detail mã đầu.
        kw = re.sub(r'(tra cứu|tìm|cho tôi|thông tin|tài liệu|sản phẩm|về|part|phụ tùng)', '',
                    question, flags=re.I).strip()
        if len(kw) >= 2:
            res = apc.results(_safe_get(user, '/api/v1/catalog/parts/', {'search': kw}))
            if res:
                part = _api_part(user, res[0].get('tokin_part_no')) or res[0]
    if not part:
        return ("Không tìm thấy phụ tùng khớp. Cho mình **mã** (VD 001002) hoặc "
                "**tên** chính xác hơn nhé.")

    spec = []
    if part.get('wire_size_mm'):  spec.append(f"dây {part['wire_size_mm']}mm")
    if part.get('thread_type'):   spec.append(f"ren {part['thread_type']}")
    if part.get('material'):      spec.append(part['material'])
    spec_s = ", ".join(spec) or "—"
    price = _vnd(part['price_vnd']) if part.get('price_vnd') else "liên hệ"
    torches = part.get('applicable_torches') or part.get('torch_models') or []
    t_s = (", ".join(map(str, torches[:6])) + ("…" if len(torches) > 6 else "")) if torches else "—"
    return (f"**{part.get('display_name_vi')}** (`{part.get('tokin_part_no')}`)\n"
            f"• Nhóm: {part.get('category') or '—'} | Spec: {spec_s}\n"
            f"• Giá: **{price}** / {part.get('price_unit')}\n"
            f"• Dùng cho súng: {t_s}")


# ── Tool đọc sâu: đơn của KH, tồn của mã (QUA API) ──────────────────────────
def tool_customer_orders(user, customer_name: str) -> str:
    if not customer_name:
        return "Cần tên khách hàng. VD: \"đơn của Công ty ABC\"."
    cust = _api_customer(user, customer_name)
    if not cust:
        return f"Không tìm thấy khách hàng \"{customer_name}\" (trong phạm vi của bạn)."
    rows = apc.results(_safe_get(user, '/api/v1/sales/orders/',
                                 {'customer': cust['id'], 'ordering': '-issued_date'}))[:8]
    if not rows:
        return f"**{cust['name']}** chưa có đơn hàng nào."
    lines = []
    for o in rows:
        debt = _num(o.get('total_vnd')) - _num(o.get('paid_vnd'))
        tail = f" — còn nợ {_vnd(debt)}" if debt > 0 else " — đã thanh toán"
        status = o.get('status_display') or o.get('status') or ''
        lines.append(f"• {o.get('code')} ({status}): {_vnd(o.get('total_vnd'))}{tail}")
    return f"Đơn hàng của **{cust['name']}** (gần nhất):\n" + "\n".join(lines)


def tool_stock_lookup(user, question: str) -> str:
    codes = re.findall(r'\b[0-9]{4,}[A-Za-z0-9\-]*\b', question)
    part = None
    for c in codes:
        part = _api_part(user, c)
        if part:
            break
    if not part:
        return "Cho mình **mã phụ tùng** để tra tồn. VD: \"tồn 001002\"."
    rows = apc.results(_safe_get(user, '/api/v1/wms/inventory/',
                                 {'part': part['tokin_part_no']}))[:20]
    total = sum(r.get('qty_on_hand') or 0 for r in rows)
    name, pn = part.get('display_name_vi'), part.get('tokin_part_no')
    if total == 0:
        return f"**{name}** (`{pn}`): hết hàng (tồn 0)."
    by = "\n".join(f"• {r.get('bin_code')}: {r.get('qty_on_hand')}" for r in rows if r.get('qty_on_hand'))
    return f"Tồn **{name}** (`{pn}`): **{total}**\n{by}"


# ── Tool tra cứu kỹ thuật: lắp đặt/sửa chữa · tương thích · bộ tiêu hao ──────
_PROC_STOP = {'cách', 'cach', 'thay', 'lắp', 'lap', 'đặt', 'dat', 'quy', 'trình', 'trinh',
              'hướng', 'huong', 'dẫn', 'dan', 'cho', 'là', 'la', 'gì', 'gi', 'của', 'cua',
              'bao', 'nhiêu', 'nhieu', 'mm', 'như', 'nhu', 'thế', 'the', 'nào', 'nao', 'và', 'va'}


def tool_procedure(user, question: str) -> str:
    """Tra cứu lắp đặt/sửa chữa/thông số kỹ thuật từ ProcedureQA — QUA API."""
    toks = [t for t in re.split(r'[\s,?]+', question.lower())
            if t and t not in _PROC_STOP and len(t) >= 2]
    if not toks:
        return "Anh/chị cho em từ khóa (liner, tip, nozzle, torque, tên súng...) để tra hướng dẫn ạ."
    rows = apc.results(_safe_get(user, '/api/v1/catalog/procedures/', {'q': ' '.join(toks)}))
    if not rows:
        return ("Em chưa thấy hướng dẫn khớp. Thử từ khóa cụ thể hơn (liner, tip, nozzle, "
                "torque, model súng...) ạ.")
    rows.sort(key=lambda r: sum(t in (r.get('question', '') + r.get('answer', '')).lower()
                                for t in toks), reverse=True)
    out = [f"**{r.get('question')}**\n{r.get('answer')}" for r in rows[:3]]
    return "\n\n".join(out)


def tool_compatibility(user, question: str) -> str:
    """Đồ đi kèm/tương thích với 1 mã — QUA API (CompatibilityEdge)."""
    codes = re.findall(r'\b\d{6}\b', question)
    if not codes:
        return "Anh/chị cho em **mã sản phẩm** (6 số, vd 001002) để tra đồ đi kèm ạ."
    pn = codes[0]
    d = _safe_get(user, '/api/v1/catalog/parts/compatibility/', {'part': pn}) or {}
    edges = d.get('results') or []
    if not edges:
        return f"Mã {pn} chưa có dữ liệu đồ đi kèm trong hệ thống."
    lines = [f"• {e['to_part']} — {e.get('name', '')}"
             + (" *(bắt buộc)*" if e.get('is_mandatory') else "") for e in edges]
    return f"Đồ đi kèm / tương thích với **{pn}**:\n" + "\n".join(lines)


def tool_consumable_set(user, question: str) -> str:
    """Bộ vật tư tiêu hao cho 1 súng hàn — QUA API (ConsumableSet)."""
    m = re.search(r'\b([A-Za-z]{2,}-?\d{2,}[A-Za-z0-9]*)\b', question)
    model = (m.group(1).upper() if m else '')
    d = _safe_get(user, '/api/v1/catalog/torches/consumable-set/', {'model': model}) or {}
    if not d.get('matched'):
        avail = d.get('available') or []
        names = "\n".join(f"• {s['set_id']}: {s.get('name', '')}" for s in avail)
        tail = f"\nMột số bộ có sẵn:\n{names}" if names else ""
        return "Anh/chị cho em **model súng hàn** (vd TK-308RR) để tra bộ tiêu hao ạ." + tail
    items = d.get('items') or []
    lines = [f"• {it['part_no']} — {it.get('note') or it.get('part_role')}"
             + (f" ×{it['default_quantity']}" if (it.get('default_quantity') or 0) > 1 else '')
             + (" *(bắt buộc)*" if it.get('is_mandatory') else '') for it in items]
    return f"**{d.get('name')}**:\n" + "\n".join(lines)


def answer(question: str, user, history=None) -> str:
    """Điểm vào chính: yêu cầu + user (+ lịch sử hội thoại) → trả lời/hành động.
    history giúp hiểu câu nối tiếp (vd trả lời tên KH cho yêu cầu báo giá đang dở)."""
    from apps.accounts.roles import can_use_intent, role_of

    role = role_of(user)
    # HYBRID: PLANNER function-calling (role-scoped, trích xuất chắc) là chính; khi planner
    # không gọi tool (lỗi mạng / ngoài phạm vi) → fallback BỘ TỪ KHÓA (không gọi LLM lần 2).
    intent = _fc_planner(question, role, history) or _keyword_intent(question)
    name = intent.get('intent', 'unknown')

    # Gate role theo intent (trừ unknown — chỉ trả hướng dẫn)
    if name != 'unknown' and not can_use_intent(role, name):
        return (f"Xin lỗi, vai trò **{role}** không có quyền dùng chức năng này. "
                f"Liên hệ quản lý nếu cần.")

    # XÁC NHẬN trước khi GHI: chỉ ghi khi planner báo confirm + câu có từ khẳng định
    # (2 lớp → không bao giờ ghi lần đầu / khi chưa "ok"). Fallback luôn = xem trước.
    _AFFIRM = ('ok', 'oke', 'okay', 'đồng ý', 'dong y', 'xác nhận', 'xac nhan', 'ghi đi',
               'ghi di', 'tạo đi', 'tao di', 'lập luôn', 'lap luon', 'đúng rồi', 'dung roi',
               'chốt', 'chot', 'yes', 'duyệt')
    _conf = bool(intent.get('confirm')) and any(a in question.lower() for a in _AFFIRM)

    if name == 'revenue':
        return tool_revenue(user, intent.get('period') or 'month')
    if name == 'customer_debt':
        cust_name = intent.get('customer_name') or _detect_customer(user, question)
        return tool_customer_debt(user, cust_name or None)
    if name == 'top_customers':
        return tool_top_customers(user)
    if name == 'dormant_customers':
        return tool_dormant_customers(user, int(intent.get('months') or 3))
    if name == 'reorder_suggestion':
        return tool_reorder(user)
    if name == 'slow_moving':
        return tool_slow_moving(user)
    if name == 'ceo_report':
        return tool_ceo_report(user)
    if name == 'evaluate_plan':
        return tool_evaluate_plan(user)
    if name == 'create_lead':
        lead_name = (intent.get('lead_name') or '').strip()
        company = (intent.get('company') or '').strip()
        phone = (intent.get('phone') or '').strip() or _parse_phone(question)
        if not lead_name:
            # Fallback từ câu: bỏ từ khóa, lấy phần tên trước dấu phẩy.
            raw = re.sub(r'(tạo|tao|thêm|them|ghi)\s+lead|lead\s+mới|khách\s+tiềm\s+năng|khach\s+tiem\s+nang|ghi\s+lead',
                         '', question, flags=re.I)
            mcom = re.search(r'(?:công ty|cong ty|cty)\s+([\w\sÀ-ỹ.]+)', raw, re.I)
            if mcom and not company:
                company = mcom.group(1).split(',')[0].strip()
                raw = raw.replace(mcom.group(0), '')
            raw = re.sub(r'\b0\d{8,10}\b', '', raw.replace('.', ''))
            lead_name = raw.strip(' ,:-') or ''
        return tool_create_lead(user, lead_name, company, phone, confirm=_conf)
    if name == 'create_quote':
        cust_name = intent.get('customer_name') or _detect_customer(user, question)
        items = intent.get('items') or _parse_items(question)
        return tool_create_quote(user, cust_name, items, confirm=_conf)
    if name == 'create_contract':
        cust_name = intent.get('customer_name') or _detect_customer(user, question)
        m = re.search(r'\bBG[-\s]?(\d+)\b', question, re.I)
        quote_code = f"BG-{int(m.group(1)):04d}" if m else (intent.get('quote_code') or '')
        return tool_create_contract(user, cust_name, quote_code, confirm=_conf)
    if name == 'wms_inbound':
        items = intent.get('items') or _parse_items(question)
        return tool_wms_inbound(user, items, confirm=_conf)
    if name == 'wms_outbound':
        items = intent.get('items') or _parse_items(question)
        cust_name = intent.get('customer_name') or _detect_customer(user, question)
        return tool_wms_outbound(user, items, cust_name, confirm=_conf)
    if name == 'customer_orders':
        cust_name = intent.get('customer_name') or _detect_customer(user, question)
        return tool_customer_orders(user, cust_name)
    if name == 'stock_lookup':
        return tool_stock_lookup(user, question)
    if name == 'procedure':
        return tool_procedure(user, question)
    if name == 'compatibility':
        return tool_compatibility(user, question)
    if name == 'consumable_set':
        return tool_consumable_set(user, question)
    if name == 'lookup_doc':
        return tool_lookup_doc(user, question)
    if name == 'software_help':
        return tool_software_help(question, role)

    return ("Em là **trợ lý nội bộ Tokinarc**. Tùy quyền của anh/chị, em có thể: "
            "**làm báo giá** (VD: \"làm báo giá cho Công ty ABC: 5 x 001002\"), "
            "xem **doanh thu**/**công nợ**/**top khách hàng**, **báo cáo CEO**, "
            "**đánh giá kế hoạch** (pipeline), hoặc **hướng dẫn dùng phần mềm** "
            "(VD: \"làm sao tạo báo giá\", \"vào đâu xem công nợ\"). Anh/chị nói rõ yêu cầu nhé.")


# ── Executive summary (tổng hợp liên phòng ban) ─────────────────────────────
def _template_summary(m: dict) -> str:
    """Tóm tắt mặc định (không cần LLM) — ghép câu từ số liệu thật."""
    lines = [
        "**Tổng quan**",
        f"Doanh thu tháng này đạt {_vnd(m['revenue_month'])} (đã thu {_vnd(m['collected_month'])}). "
        f"Công nợ phải thu {_vnd(m['debt_total'])}, trong đó quá hạn {_vnd(m['overdue'])}. "
        f"Pipeline weighted {_vnd(m['pipeline_weighted'])} từ {m['open_opportunities']} cơ hội đang mở.",
        "",
        "**Kinh doanh (Sales/CRM)**",
        f"• {m['customers']} khách hàng, {m['open_leads']} lead đang theo.",
        f"• Khách hàng đóng góp lớn nhất: {m['top_customer'] or '—'} ({_vnd(m['top_customer_revenue'])}).",
        f"• {m['dormant_customers']} khách chưa mua >3 tháng (nguy cơ rời).",
        "",
        "**Dịch vụ**",
        f"• {m['open_tickets']} ticket đang mở" + (f", {m['urgent_tickets']} khẩn." if m['urgent_tickets'] else "."),
        "",
        "**Kho vận**",
        f"• Giá trị tồn kho {_vnd(m['inventory_value'])} ({m['sku_count']} SKU), {m['low_stock']} mặt hàng sắp hết.",
    ]
    hd = m.get('hoat_dong')
    if hd:
        wh = hd.get('hoat_dong_kho', {})
        lines += [
            "",
            "**Hoạt động nổi bật** (30 ngày)",
            f"• Sales: {hd['so_cuoc_gap']} cuộc gặp ({hd['cuoc_gap_co_ghi_am']} có ghi âm), "
            f"{hd['so_cuoc_goi_email']} cuộc gọi/email ({hd['goi_email_co_ghi_am']} có ghi âm).",
        ]
        for r in hd.get('recap_cuoc_gap', [])[:5]:
            tail = f" → {r['viec_tiep']}" if r['viec_tiep'] else ""
            mic = " 🎙" if r['co_ghi_am'] else ""
            lines.append(f"  - {r['kh']} ({r['ngay']}){mic}: {r['recap']}{tail}")
        lines.append(
            f"• Kho: {wh.get('phieu_nhap', 0)} lượt nhập, {wh.get('phieu_xuat', 0)} lượt xuất, "
            f"{wh.get('dieu_chinh_ton', 0)} điều chỉnh tồn"
            + (f", {wh['kiem_ke']} phiên kiểm kê." if 'kiem_ke' in wh else "."))
    return "\n".join(lines)


def _llm_summary(m: dict) -> str | None:
    """Nhờ Gemini viết tóm tắt điều hành từ metrics thật (không bịa số)."""
    key = _gemini_key()
    if not key:
        return None
    model = _gemini_model()
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}"
    prompt = (
        "Bạn là trợ lý điều hành. Dưới đây là DỮ LIỆU THẬT của công ty phân phối súng hàn "
        "Tokinarc theo từng phòng ban: phần SỐ LIỆU (VND) và phần 'hoat_dong' gồm RECAP các "
        "cuộc gặp/gọi khách gần đây + hoạt động kho. Hãy viết bản TÓM TẮT ĐIỀU HÀNH bằng "
        "tiếng Việt cho Giám đốc, gồm 5 mục in đậm: **Tổng quan**, **Điểm tích cực**, "
        "**Cần chú ý**, **Hoạt động nổi bật**, **Khuyến nghị tuần này**.\n"
        "- Mục 'Hoạt động nổi bật': tóm tắt NỘI DUNG các cuộc gặp/gọi (khách nói gì, nhu cầu, "
        "phàn nàn, việc cần follow-up) dựa trên recap, nêu rõ tên khách; nêu hoạt động kho "
        "(nhập/xuất/kiểm kê/điều chỉnh) và số cuộc CÓ GHI ÂM.\n"
        "CHỈ dùng đúng dữ liệu cho sẵn, KHÔNG bịa số/nội dung. Súc tích.\n\nDỮ LIỆU:\n"
        + json.dumps(m, ensure_ascii=False)
    )
    body = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.3, "maxOutputTokens": 1200,
                             "thinkingConfig": {"thinkingBudget": 0}},
    }).encode('utf-8')
    try:
        req = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read())
        return data['candidates'][0]['content']['parts'][0]['text'].strip()
    except (urllib.error.URLError, KeyError, IndexError, ValueError, TimeoutError):
        return None


_ASSIST_FILE_PROMPT = (
    "Bạn là trợ lý kỹ thuật NỘI BỘ của Autoss (phân phối súng hàn Tokinarc). "
    "Nhân viên đính kèm 1 file (ảnh / PDF / Excel). Hãy PHÂN TÍCH nội dung và trả lời bằng "
    "tiếng Việt, súc tích, có cấu trúc.\n"
    "- Nếu là ẢNH linh kiện/súng hàn: nhận diện loại (béc/chụp/liner/collet/súng...), nêu đặc "
    "điểm thấy được, gợi ý mã Tokin tương đương nếu nhận ra; nếu thấy hư hỏng → nêu cách xử lý/thay.\n"
    "- Nếu là PDF/Excel/bảng: tóm tắt nội dung chính và trả lời đúng yêu cầu của nhân viên.\n"
    "KHÔNG bịa thông tin. Nếu không chắc → nói rõ cần thêm thông tin gì."
)


def _excel_to_text(file_bytes: bytes, max_rows: int = 200) -> str:
    """Đọc .xlsx → text gọn cho LLM (openpyxl không gửi file lên Gemini)."""
    import io as _io
    try:
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return "(không đọc được Excel)"
    out = []
    for ws in wb.worksheets[:5]:
        out.append(f"[Sheet: {ws.title}]")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                out.append("…(còn nữa)")
                break
            out.append(' | '.join('' if c is None else str(c) for c in row))
    return '\n'.join(out)[:20000]


def analyze_attachment(question: str, file_bytes: bytes, mime: str, filename: str) -> str:
    """Phân tích ẢNH / PDF / EXCEL đính kèm bằng Gemini đa phương thức."""
    import base64 as _b64

    key = _gemini_key()
    if not key:
        return "Chưa cấu hình Gemini (GEMINI_API_KEY) — không phân tích được file."
    model = _gemini_model()
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}"
    q = (question or '').strip() or "Phân tích nội dung file này giúp tôi."
    mime = (mime or '').lower()
    fname = (filename or '').lower()

    parts: list[dict] = [{"text": _ASSIST_FILE_PROMPT + f"\n\nYÊU CẦU CỦA NHÂN VIÊN: {q}"}]
    if mime.startswith('image/') or mime == 'application/pdf':
        parts.append({"inline_data": {"mime_type": mime, "data": _b64.b64encode(file_bytes).decode()}})
    elif 'spreadsheet' in mime or 'excel' in mime or fname.endswith(('.xlsx', '.xls')):
        parts[0]["text"] += "\n\nNỘI DUNG EXCEL:\n" + _excel_to_text(file_bytes)
    else:
        try:
            parts[0]["text"] += "\n\nNỘI DUNG FILE:\n" + file_bytes.decode('utf-8', errors='ignore')[:20000]
        except Exception:  # noqa: BLE001
            return "Định dạng file chưa hỗ trợ. Em hỗ trợ ảnh (jpg/png), PDF, Excel, văn bản."

    body = json.dumps({
        "contents": [{"parts": parts}],
        "generationConfig": {"temperature": 0.3, "maxOutputTokens": 1400,
                             "thinkingConfig": {"thinkingBudget": 0}},
    }).encode('utf-8')
    try:
        req = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'})
        with urllib.request.urlopen(req, timeout=45) as resp:
            data = json.loads(resp.read())
        return data['candidates'][0]['content']['parts'][0]['text'].strip()
    except (urllib.error.URLError, KeyError, IndexError, ValueError, TimeoutError) as e:
        return f"Em chưa phân tích được file (lỗi gọi AI): {e}"


def executive_summary(user) -> dict:
    """Tóm tắt toàn bộ hoạt động phòng ban. Số liệu THẬT lấy QUA API (executive-metrics),
    lời do LLM (fallback template). Bot không truy vấn DB trực tiếp."""
    m = _safe_get(user, '/api/v1/analytics/executive-metrics/') or {}
    if not m:
        return {'summary': 'Chưa lấy được số liệu điều hành (kiểm tra quyền hoặc kết nối).',
                'metrics': {}, 'generated_by': 'none'}
    ai = _llm_summary(m)
    return {
        'summary': ai or _template_summary(m),
        'metrics': m,
        'generated_by': 'ai' if ai else 'template',
    }
