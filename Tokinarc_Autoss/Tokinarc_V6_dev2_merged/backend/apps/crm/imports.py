"""
Tokinarc V6 — apps/crm/imports.py

Import dữ liệu KHÁCH HÀNG cũ (trước khi có phần mềm) từ Excel/CSV.
  - POST /api/v1/crm/customers/import/         (file=..., dry_run=1 để xem trước)
  - GET  /api/v1/crm/customers/import-template/ (tải file Excel mẫu)

Quy tắc:
  - Chỉ manager/CEO/admin được import (nghiệp vụ nhạy cảm, ghi hàng loạt).
  - Dedup theo `code`: KH đã tồn tại → BỎ QUA (không ghi đè).
  - dry_run: chỉ kiểm tra + trả thống kê/lỗi, KHÔNG ghi DB.
  - Mỗi dòng = 1 KH + (tùy chọn) 1 người liên hệ chính.
"""
from __future__ import annotations

import csv
import io

from django.db import transaction
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.roles import is_manager

from .models import (
    Contact, ContactChannel, Contract, ContractStatus, Customer,
    CustomerSegment, CustomerStatus, Lead, LeadStatus,
)

# Cột nhận diện được (header — thường/không dấu đều chấp nhận qua _norm).
COLUMNS = ['code', 'name', 'segment', 'region', 'tax_code', 'status', 'notes',
           'owner_username', 'contact_name', 'contact_phone', 'contact_email',
           'contact_title']

# Map nhãn tiếng Việt + value → value chuẩn
_SEGMENT = {**{c.value: c.value for c in CustomerSegment},
            **{c.label.lower(): c.value for c in CustomerSegment},
            'nha may': 'factory', 'nhà máy': 'factory', 'đại lý': 'dealer',
            'dai ly': 'dealer', 'đóng tàu': 'shipyard', 'dong tau': 'shipyard'}
_STATUS = {**{c.value: c.value for c in CustomerStatus},
           **{c.label.lower(): c.value for c in CustomerStatus},
           'tiềm năng': 'potential', 'tiem nang': 'potential', 'mới': 'new'}


def _norm(s: str) -> str:
    return (s or '').strip().lower().replace(' ', '_').replace('-', '_')


def _parse_file(f) -> list[dict]:
    """Đọc Excel (.xlsx) hoặc CSV → list dict {header: value}."""
    name = (getattr(f, 'name', '') or '').lower()
    if name.endswith('.csv'):
        text = f.read().decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(text))
        return [{_norm(k): (v or '').strip() for k, v in row.items()} for row in reader]
    # Excel
    from openpyxl import load_workbook
    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [_norm(str(h)) if h is not None else '' for h in next(rows)]
    except StopIteration:
        return []
    out = []
    for r in rows:
        if r is None or all(c is None for c in r):
            continue
        out.append({headers[i]: (str(c).strip() if c is not None else '')
                    for i, c in enumerate(r) if i < len(headers)})
    return out


def _resolve_owner(username: str, default_user):
    if not username:
        return default_user
    from apps.accounts.models import User
    return User.objects.filter(username=username).first() or default_user


def validate_and_build(rows: list[dict], user):
    """Trả (valid_items, errors, skipped). Không ghi DB."""
    valid, errors, skipped = [], [], 0
    seen_codes = set()
    existing = set(Customer.all_objects.values_list('code', flat=True)) \
        if hasattr(Customer, 'all_objects') else set(Customer.objects.values_list('code', flat=True))

    for i, row in enumerate(rows, start=2):   # dòng 1 = header
        code = (row.get('code') or '').strip()
        name = (row.get('name') or '').strip()
        if not code and not name:
            continue   # dòng trống
        if not code:
            errors.append({'row': i, 'message': 'Thiếu mã KH (code).'}); continue
        if not code.upper().startswith('KH'):
            errors.append({'row': i, 'message': f'Mã "{code}" phải bắt đầu bằng KH.'}); continue
        if not name:
            errors.append({'row': i, 'message': f'KH {code}: thiếu tên.'}); continue
        if code in existing or code in seen_codes:
            skipped += 1; continue
        seen_codes.add(code)

        valid.append({
            'code': code, 'name': name,
            'segment': _SEGMENT.get((row.get('segment') or '').strip().lower(), 'other'),
            'status':  _STATUS.get((row.get('status') or '').strip().lower(), 'new'),
            'region':  (row.get('region') or '').strip(),
            'tax_code': (row.get('tax_code') or '').strip(),
            'notes':   (row.get('notes') or '').strip(),
            'owner_username': (row.get('owner_username') or '').strip(),
            'contact_name': (row.get('contact_name') or '').strip(),
            'contact_phone': (row.get('contact_phone') or '').strip(),
            'contact_email': (row.get('contact_email') or '').strip(),
            'contact_title': (row.get('contact_title') or '').strip(),
        })
    return valid, errors, skipped


class CustomerImportView(APIView):
    """Import KH cũ từ Excel/CSV. ?dry_run=1 để xem trước (không ghi)."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if not is_manager(request.user):
            return Response({'detail': 'Chỉ quản lý/CEO/admin được import dữ liệu.'}, status=403)
        f = request.FILES.get('file')
        if not f:
            return Response({'detail': 'Thiếu file.'}, status=400)
        try:
            rows = _parse_file(f)
        except Exception as e:   # noqa: BLE001 — báo lỗi đọc file thân thiện
            return Response({'detail': f'Không đọc được file: {e}'}, status=400)

        valid, errors, skipped = validate_and_build(rows, request.user)
        dry = str(request.query_params.get('dry_run', '')).lower() in ('1', 'true', 'yes')

        if dry:
            return Response({
                'dry_run': True, 'total_rows': len(rows),
                'will_create': len(valid), 'skipped_existing': skipped,
                'errors': errors,
                'preview': [{'code': v['code'], 'name': v['name'], 'segment': v['segment']}
                            for v in valid[:10]],
            })

        created = 0
        with transaction.atomic():
            for v in valid:
                owner = _resolve_owner(v['owner_username'], request.user)
                cust = Customer.objects.create(
                    code=v['code'], name=v['name'], segment=v['segment'],
                    status=v['status'], region=v['region'], tax_code=v['tax_code'],
                    notes=v['notes'], owner=owner,
                )
                if v['contact_name']:
                    Contact.objects.create(
                        customer=cust, full_name=v['contact_name'],
                        phone=v['contact_phone'], email=v['contact_email'],
                        title=v['contact_title'], is_primary=True,
                        preferred_channel=ContactChannel.ZALO,
                    )
                created += 1
        return Response({'dry_run': False, 'created': created,
                         'skipped_existing': skipped, 'errors': errors})


class CustomerImportTemplateView(APIView):
    """Tải file Excel mẫu (đúng cột) để điền dữ liệu KH cũ."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.http import HttpResponse
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = 'KhachHang'
        ws.append(COLUMNS)
        ws.append(['KH-1001', 'Công ty TNHH ABC', 'factory', 'HCM', '0312345678',
                   'potential', 'KH nhập từ dữ liệu cũ', '', 'Nguyễn Văn A',
                   '0901234567', 'a@abc.vn', 'Giám đốc'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="mau_import_khach_hang.xlsx"'
        return resp


# ─────────────────────────────────────────────────────────────────────────────
# Phase 2 — Import Lead / Hợp đồng / Đơn hàng cũ (registry chung)
# ─────────────────────────────────────────────────────────────────────────────
import re   # noqa: E402


def _to_int(v) -> int:
    """'1.200.000' / '1,200,000' / '1200000' → 1200000."""
    digits = re.sub(r'[^\d]', '', str(v or ''))
    return int(digits) if digits else 0


def _to_date(v):
    from datetime import datetime
    s = str(v or '').strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _choice_map(choices) -> dict:
    m = {}
    for c in choices:
        m[c.value] = c.value
        m[c.label.lower()] = c.value
    return m


_LEAD_STATUS = _choice_map(LeadStatus)
_CONTRACT_STATUS = _choice_map(ContractStatus)


def _customer_by_code():
    return {c.code: c for c in Customer.objects.all()}


# ── Lead ────────────────────────────────────────────────────────────────────
def _validate_leads(rows, user):
    valid, errors, skipped = [], [], 0
    existing = {(n or '').lower() + '|' + (p or '')
                for n, p in Lead.objects.values_list('name', 'phone')}
    seen = set()
    for i, row in enumerate(rows, start=2):
        name = (row.get('name') or '').strip()
        if not name and not any(row.values()):
            continue
        if not name:
            errors.append({'row': i, 'message': 'Thiếu tên lead.'}); continue
        phone = (row.get('phone') or '').strip()
        key = name.lower() + '|' + phone
        if key in existing or key in seen:
            skipped += 1; continue
        seen.add(key)
        valid.append({
            'name': name, 'company': (row.get('company') or '').strip(),
            'phone': phone, 'email': (row.get('email') or '').strip(),
            'source': (row.get('source') or '').strip(),
            'status': _LEAD_STATUS.get((row.get('status') or '').strip().lower(), 'new'),
            'score': _to_int(row.get('score')),
            'owner_username': (row.get('owner_username') or '').strip(),
            'notes': (row.get('notes') or '').strip(),
        })
    return valid, errors, skipped


def _create_leads(valid, user):
    n = 0
    for v in valid:
        Lead.objects.create(
            name=v['name'], company=v['company'], phone=v['phone'], email=v['email'],
            source=v['source'], status=v['status'], score=v['score'],
            owner=_resolve_owner(v['owner_username'], user), notes=v['notes'])
        n += 1
    return n


# ── Hợp đồng ────────────────────────────────────────────────────────────────
def _validate_contracts(rows, user):
    valid, errors, skipped = [], [], 0
    cmap = _customer_by_code()
    existing = set(Contract.objects.values_list('code', flat=True))
    seen = set()
    for i, row in enumerate(rows, start=2):
        code = (row.get('code') or '').strip()
        ccode = (row.get('customer_code') or '').strip()
        if not code and not ccode and not any(row.values()):
            continue
        if not code:
            errors.append({'row': i, 'message': 'Thiếu mã hợp đồng (code).'}); continue
        if code in existing or code in seen:
            skipped += 1; continue
        cust = cmap.get(ccode)
        if not cust:
            errors.append({'row': i, 'message': f'HĐ {code}: không tìm thấy KH mã "{ccode}".'}); continue
        seen.add(code)
        valid.append({
            'code': code, 'customer': cust,
            'title': (row.get('title') or '').strip(),
            'value_vnd': _to_int(row.get('value_vnd')),
            'paid_vnd': _to_int(row.get('paid_vnd')),
            'status': _CONTRACT_STATUS.get((row.get('status') or '').strip().lower(), 'active'),
            'start_date': _to_date(row.get('start_date')),
            'end_date': _to_date(row.get('end_date')),
            'owner_username': (row.get('owner_username') or '').strip(),
            'notes': (row.get('notes') or '').strip(),
        })
    return valid, errors, skipped


def _create_contracts(valid, user):
    n = 0
    for v in valid:
        Contract.objects.create(
            code=v['code'], customer=v['customer'], title=v['title'],
            value_vnd=v['value_vnd'], paid_vnd=v['paid_vnd'], status=v['status'],
            start_date=v['start_date'], end_date=v['end_date'],
            owner=_resolve_owner(v['owner_username'], user), notes=v['notes'])
        n += 1
    return n


# ── Đơn hàng ────────────────────────────────────────────────────────────────
def _validate_orders(rows, user):
    from apps.sales.models import OrderStatus
    status_map = _choice_map(OrderStatus)
    valid, errors, skipped = [], [], 0
    cmap = _customer_by_code()
    from apps.sales.models import SalesOrder
    existing = set(SalesOrder.objects.values_list('code', flat=True))
    seen = set()
    for i, row in enumerate(rows, start=2):
        code = (row.get('code') or '').strip()
        ccode = (row.get('customer_code') or '').strip()
        if not code and not ccode and not any(row.values()):
            continue
        if not code:
            errors.append({'row': i, 'message': 'Thiếu mã đơn (code).'}); continue
        if code in existing or code in seen:
            skipped += 1; continue
        cust = cmap.get(ccode)
        if not cust:
            errors.append({'row': i, 'message': f'Đơn {code}: không tìm thấy KH mã "{ccode}".'}); continue
        issued = _to_date(row.get('issued_date'))
        if not issued:
            errors.append({'row': i, 'message': f'Đơn {code}: thiếu/sai ngày (issued_date).'}); continue
        total, paid = _to_int(row.get('total_vnd')), _to_int(row.get('paid_vnd'))
        if paid > total:
            errors.append({'row': i, 'message': f'Đơn {code}: đã thu ({paid}) > tổng ({total}).'}); continue
        seen.add(code)
        valid.append({
            'code': code, 'customer': cust, 'issued_date': issued,
            'total_vnd': total, 'paid_vnd': paid,
            'status': status_map.get((row.get('status') or '').strip().lower(), 'completed'),
            'owner_username': (row.get('owner_username') or '').strip(),
            'notes': (row.get('notes') or '').strip(),
        })
    return valid, errors, skipped


def _create_orders(valid, user):
    from apps.sales.models import SalesOrder
    n = 0
    for v in valid:
        SalesOrder.objects.create(
            code=v['code'], customer=v['customer'], issued_date=v['issued_date'],
            total_vnd=v['total_vnd'], paid_vnd=v['paid_vnd'], status=v['status'],
            owner=_resolve_owner(v['owner_username'], user), notes=v['notes'])
        n += 1
    return n


# Registry: entity → cấu hình import
ENTITY_SPECS = {
    'leads': {
        'label': 'Lead',
        'columns': ['name', 'company', 'phone', 'email', 'source', 'status',
                    'score', 'owner_username', 'notes'],
        'example': ['Anh Hùng', 'Cong ty XYZ', '0907', 'hung@xyz.vn', 'web',
                    'new', '50', '', 'Lead nhập từ dữ liệu cũ'],
        'validate': _validate_leads, 'create': _create_leads,
    },
    'contracts': {
        'label': 'Hợp đồng',
        'columns': ['code', 'customer_code', 'title', 'value_vnd', 'paid_vnd',
                    'status', 'start_date', 'end_date', 'owner_username', 'notes'],
        'example': ['HD-2024-001', 'KH-1001', 'Hợp đồng khung 2024', '500000000',
                    '200000000', 'active', '2024-01-15', '2024-12-31', '', ''],
        'validate': _validate_contracts, 'create': _create_contracts,
    },
    'orders': {
        'label': 'Đơn hàng',
        'columns': ['code', 'customer_code', 'issued_date', 'total_vnd', 'paid_vnd',
                    'status', 'owner_username', 'notes'],
        'example': ['DH-2024-001', 'KH-1001', '2024-03-20', '120000000',
                    '120000000', 'completed', '', ''],
        'validate': _validate_orders, 'create': _create_orders,
    },
}


class EntityImportView(APIView):
    """Import Lead/Hợp đồng/Đơn hàng cũ. POST /crm/import/<entity>/?dry_run=1."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, entity):
        spec = ENTITY_SPECS.get(entity)
        if not spec:
            return Response({'detail': f'Không hỗ trợ import "{entity}".'}, status=404)
        if not is_manager(request.user):
            return Response({'detail': 'Chỉ quản lý/CEO/admin được import dữ liệu.'}, status=403)
        f = request.FILES.get('file')
        if not f:
            return Response({'detail': 'Thiếu file.'}, status=400)
        try:
            rows = _parse_file(f)
        except Exception as e:   # noqa: BLE001
            return Response({'detail': f'Không đọc được file: {e}'}, status=400)

        valid, errors, skipped = spec['validate'](rows, request.user)
        dry = str(request.query_params.get('dry_run', '')).lower() in ('1', 'true', 'yes')
        if dry:
            return Response({'dry_run': True, 'entity': entity, 'total_rows': len(rows),
                             'will_create': len(valid), 'skipped_existing': skipped,
                             'errors': errors})
        with transaction.atomic():
            created = spec['create'](valid, request.user)
        return Response({'dry_run': False, 'entity': entity, 'created': created,
                         'skipped_existing': skipped, 'errors': errors})


class EntityImportTemplateView(APIView):
    """Tải Excel mẫu cho Lead/Hợp đồng/Đơn hàng. GET /crm/import/<entity>/template/."""
    permission_classes = [IsAuthenticated]

    def get(self, request, entity):
        spec = ENTITY_SPECS.get(entity)
        if not spec:
            return Response({'detail': f'Không hỗ trợ "{entity}".'}, status=404)
        from django.http import HttpResponse
        from openpyxl import Workbook

        wb = Workbook(); ws = wb.active; ws.title = entity
        ws.append(spec['columns']); ws.append(spec['example'])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="mau_import_{entity}.xlsx"'
        return resp
